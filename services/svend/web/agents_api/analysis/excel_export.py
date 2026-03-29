"""Excel export for analysis results.

Generates .xlsx files from DSWResult records. Includes:
- Summary sheet with analysis metadata and text summary
- Statistics sheet with key metrics as a table
- Data sheet with the source data (if available)

CR: bb579300
"""

import io
import json
import logging
import re

from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods

from accounts.permissions import require_auth

from ..models import DSWResult

logger = logging.getLogger(__name__)


def _strip_html(text):
    """Remove HTML tags and color markup from summary text."""
    if not text:
        return ""
    text = re.sub(r"<<COLOR:\w+>>", "", text)
    text = re.sub(r"<</COLOR>>", "", text)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def _write_summary_sheet(wb, result_data):
    """Write summary sheet with analysis metadata and text summary."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.active
    ws.title = "Summary"

    header_font = Font(name="Calibri", size=14, bold=True)
    meta_font = Font(name="Calibri", size=11, color="444444")
    body_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # Title
    analysis_type = result_data.get("analysis_type", "")
    analysis_id = result_data.get("analysis_id", "")
    title = f"{analysis_type.upper()} — {analysis_id.replace('_', ' ').title()}"
    ws["A1"] = title
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:D1")

    # Metadata
    ws["A3"] = "Analysis Type"
    ws["B3"] = analysis_type
    ws["A4"] = "Analysis"
    ws["B4"] = analysis_id
    for row in range(3, 5):
        ws.cell(row=row, column=1).font = meta_font
        ws.cell(row=row, column=2).font = body_font

    # Config summary
    config = result_data.get("config", {})
    if config:
        ws["A6"] = "Configuration"
        ws["A6"].font = Font(name="Calibri", size=12, bold=True)
        row = 7
        for key, value in config.items():
            if key in ("data", "raw_data"):
                continue
            ws.cell(row=row, column=1, value=str(key)).font = meta_font
            display = str(value) if not isinstance(value, (list, dict)) else json.dumps(value)
            ws.cell(row=row, column=2, value=display[:200]).font = body_font
            row += 1

    # Summary text
    summary = _strip_html(result_data.get("summary", ""))
    if summary:
        summary_row = row + 1
        ws.cell(row=summary_row, column=1, value="Results Summary").font = Font(name="Calibri", size=12, bold=True)
        summary_row += 1
        # Split long summary into lines
        for line in summary.split("\n"):
            line = line.strip()
            if line:
                cell = ws.cell(row=summary_row, column=1, value=line)
                cell.font = body_font
                cell.alignment = Alignment(wrap_text=True)
                summary_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _write_statistics_sheet(wb, result_data):
    """Write statistics sheet with key metrics."""
    from openpyxl.styles import Font, PatternFill

    stats = result_data.get("statistics", {})
    if not stats:
        return

    ws = wb.create_sheet("Statistics")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A9F6E", end_color="4A9F6E", fill_type="solid")
    body_font = Font(name="Calibri", size=11)

    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["B1"].font = header_font
    ws["B1"].fill = header_fill

    row = 2
    for key, value in stats.items():
        if isinstance(value, (dict, list)):
            continue
        ws.cell(row=row, column=1, value=str(key).replace("_", " ").title()).font = body_font
        cell = ws.cell(row=row, column=2)
        if isinstance(value, float):
            cell.value = round(value, 6)
            cell.number_format = "0.####"
        else:
            cell.value = value
        cell.font = body_font
        row += 1

    # Nested stats (e.g., ANOVA tables)
    for key, value in stats.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            ws_nested = wb.create_sheet(key.replace("_", " ").title()[:31])
            headers = list(value[0].keys())
            for col, h in enumerate(headers, 1):
                cell = ws_nested.cell(row=1, column=col, value=h.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
            for r, item in enumerate(value, 2):
                for col, h in enumerate(headers, 1):
                    v = item.get(h)
                    cell = ws_nested.cell(row=r, column=col)
                    if isinstance(v, float):
                        cell.value = round(v, 6)
                        cell.number_format = "0.####"
                    else:
                        cell.value = v
                    cell.font = body_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25


def _build_xlsx(result_data):
    """Build an xlsx workbook from result data. Returns (workbook, filename)."""
    from openpyxl import Workbook

    wb = Workbook()
    _write_summary_sheet(wb, result_data)
    _write_statistics_sheet(wb, result_data)

    analysis_id = result_data.get("analysis_id", "analysis")
    filename = f"svend_{analysis_id}.xlsx"
    return wb, filename


def _xlsx_response(wb, filename):
    """Convert workbook to HttpResponse."""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_http_methods(["GET"])
@require_auth
def export_xlsx(request, result_id):
    """Export a saved analysis result as .xlsx.

    GET /api/analysis/export/xlsx/<result_id>/
    """
    try:
        result_obj = DSWResult.objects.get(id=result_id, user=request.user)
    except DSWResult.DoesNotExist:
        return JsonResponse({"error": "Result not found"}, status=404)

    try:
        result_data = json.loads(result_obj.data)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({"error": "Invalid result data"}, status=500)

    wb, filename = _build_xlsx(result_data)
    return _xlsx_response(wb, filename)


@require_http_methods(["POST"])
@require_auth
def export_xlsx_inline(request):
    """Export analysis result JSON directly as .xlsx (no saved result needed).

    POST /api/analysis/export/xlsx/ with JSON body containing the result data.
    Limited to 500KB to prevent abuse.
    """
    if len(request.body) > 512_000:
        return JsonResponse({"error": "Request too large"}, status=413)

    try:
        result_data = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    wb, filename = _build_xlsx(result_data)
    return _xlsx_response(wb, filename)
