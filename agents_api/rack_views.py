"""ForgeRack session CRUD — save/load rack configurations."""

import json
import statistics

from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_http_methods

from accounts.permissions import require_auth
from agents_api.models import RackSession

FREE_SESSION_LIMIT = 3


def _rack_queryset(user):
    """Get rack sessions accessible to the user."""
    from django.db.models import Q

    qs = RackSession.objects.filter(Q(user=user))
    if hasattr(user, "tenant") and user.tenant:
        qs = RackSession.objects.filter(Q(user=user) | Q(tenant=user.tenant))
    return qs


def _can_edit(user, session):
    if session.user_id == user.id:
        return True
    if session.tenant and hasattr(user, "tenant") and session.tenant == user.tenant:
        return True
    return False


def _check_limit(user):
    """Free users limited to FREE_SESSION_LIMIT sessions."""
    sub = getattr(user, "subscription", None)
    if sub and getattr(sub, "tier", "free") != "free":
        return True  # Paid — no limit
    count = RackSession.objects.filter(user=user).count()
    return count < FREE_SESSION_LIMIT


@require_auth
@require_http_methods(["GET"])
def list_rack_sessions(request):
    sessions = _rack_queryset(request.user).values("id", "title", "session_type", "status", "updated_at")
    return JsonResponse(
        [
            {
                "id": str(s["id"]),
                "title": s["title"],
                "session_type": s["session_type"],
                "status": s["status"],
                "updated_at": s["updated_at"].isoformat() if s["updated_at"] else None,
            }
            for s in sessions
        ],
        safe=False,
    )


@require_auth
@require_http_methods(["POST"])
def create_rack_session(request):
    if not _check_limit(request.user):
        return JsonResponse(
            {"error": f"Free tier limited to {FREE_SESSION_LIMIT} saved sessions"},
            status=403,
        )

    data = json.loads(request.body)
    session = RackSession.objects.create(
        title=data.get("title", "Untitled Rack"),
        description=data.get("description", ""),
        session_type=data.get("session_type", RackSession.SessionType.SANDBOX),
        state=data.get("state", {}),
        user=request.user,
        tenant=getattr(request.user, "tenant", None),
    )
    return JsonResponse(session.to_dict(), status=201)


@require_auth
@require_http_methods(["GET"])
def get_rack_session(request, session_id):
    session = get_object_or_404(RackSession, id=session_id)
    if not _can_edit(request.user, session):
        return JsonResponse({"error": "Not authorized"}, status=403)
    return JsonResponse(session.to_dict())


@require_auth
@require_http_methods(["POST", "PUT"])
def update_rack_session(request, session_id):
    session = get_object_or_404(RackSession, id=session_id)
    if not _can_edit(request.user, session):
        return JsonResponse({"error": "Not authorized"}, status=403)

    data = json.loads(request.body)
    if "title" in data:
        session.title = data["title"]
    if "description" in data:
        session.description = data["description"]
    if "state" in data:
        session.state = data["state"]
    if "status" in data:
        session.status = data["status"]
    if "session_type" in data:
        session.session_type = data["session_type"]
    session.save()
    return JsonResponse(session.to_dict())


@require_auth
@require_http_methods(["DELETE"])
def delete_rack_session(request, session_id):
    session = get_object_or_404(RackSession, id=session_id)
    if not _can_edit(request.user, session):
        return JsonResponse({"error": "Not authorized"}, status=403)
    session.delete()
    return JsonResponse({"status": "deleted"})


# ═══════════════════════════════════════════════════════════════
# Rack Compute — forge package bridge for client-side units
# ═══════════════════════════════════════════════════════════════

_RACK_OPS = {}


def _rack_op(name):
    """Decorator to register a rack compute operation."""

    def decorator(fn):
        _RACK_OPS[name] = fn
        return fn

    return decorator


@_rack_op("mean")
def _op_mean(d):
    return {"value": statistics.mean(d["values"])}


@_rack_op("median")
def _op_median(d):
    return {"value": statistics.median(d["values"])}


@_rack_op("stdev")
def _op_stdev(d):
    return {"value": statistics.stdev(d["values"])}


@_rack_op("descriptive")
def _op_descriptive(d):
    v = d["values"]
    return {
        "mean": statistics.mean(v),
        "median": statistics.median(v),
        "stdev": statistics.stdev(v) if len(v) > 1 else 0,
        "min": min(v),
        "max": max(v),
        "n": len(v),
        "range": max(v) - min(v),
        "sum": sum(v),
    }


@_rack_op("ttest_2sample")
def _op_ttest(d):
    from forgestat.parametric.ttest import two_sample

    result = two_sample(d["a"], d["b"])
    alpha = d.get("alpha", 0.05)
    return {
        "t": result.statistic,
        "p": result.p_value,
        "df": result.df,
        "ci_lower": getattr(result, "ci_lower", None),
        "ci_upper": getattr(result, "ci_upper", None),
        "effect_size": getattr(result, "effect_size", None),
        "significant": result.p_value < alpha,
        "mean_a": statistics.mean(d["a"]),
        "mean_b": statistics.mean(d["b"]),
        "n_a": len(d["a"]),
        "n_b": len(d["b"]),
    }


def _has_variance(vals):
    """Check if a list has non-zero variance (not all identical)."""
    if len(vals) < 2:
        return False
    return any(v != vals[0] for v in vals[1:])


@_rack_op("pearson")
def _op_pearson(d):
    x, y = d["x"], d["y"]
    if not _has_variance(x) or not _has_variance(y):
        return {"r": 0.0, "p": 1.0, "r_squared": 0.0, "n": len(x), "warning": "zero variance"}
    from forgestat.parametric.correlation import correlation

    result = correlation({"x": x, "y": y}, method="pearson")
    pair = result.pairs[0]
    return {"r": pair.r, "p": pair.p_value, "r_squared": pair.r_squared, "n": pair.n}


@_rack_op("spearman")
def _op_spearman(d):
    x, y = d["x"], d["y"]
    if not _has_variance(x) or not _has_variance(y):
        return {"rho": 0.0, "p": 1.0, "n": len(x), "warning": "zero variance"}
    from forgestat.parametric.correlation import correlation

    result = correlation({"x": x, "y": y}, method="spearman")
    pair = result.pairs[0]
    return {"rho": pair.r, "p": pair.p_value, "n": pair.n}


@_rack_op("control_chart")
def _op_control_chart(d):
    from forgespc.advanced import cusum_chart, ewma_chart, xbar_s_chart
    from forgespc.charts import (
        c_chart,
        individuals_moving_range_chart,
        np_chart,
        p_chart,
        u_chart,
        xbar_r_chart,
    )

    vals = d["values"]
    chart_type = d.get("chart_type", "imr")
    rules_mode = d.get("rules", "nelson")
    lsl = d.get("lsl")
    usl = d.get("usl")
    subgroups = d.get("subgroups")  # list of lists for X-bar charts

    # ── Dispatch by chart type ──────────────────────────────────
    if chart_type in ("xbar_r", "xbar_s") and subgroups:
        if chart_type == "xbar_r":
            r = xbar_r_chart(subgroups, usl=usl, lsl=lsl)
        else:
            r = xbar_s_chart(subgroups)
    elif chart_type == "p":
        sample_sizes = d.get("sample_sizes", [len(vals)] * len(vals))
        r = p_chart([int(v) for v in vals], [int(s) for s in sample_sizes])
    elif chart_type == "np":
        sample_size = d.get("sample_size", len(vals))
        r = np_chart([int(v) for v in vals], int(sample_size))
    elif chart_type == "c":
        r = c_chart([int(v) for v in vals])
    elif chart_type == "u":
        unit_counts = d.get("unit_counts", [1.0] * len(vals))
        r = u_chart([int(v) for v in vals], [float(u) for u in unit_counts])
    elif chart_type == "cusum":
        cr = cusum_chart(vals)
        # CUSUM has a different result shape
        ooc_indices = set(cr.signals_up + cr.signals_down)
        result = {
            "mean": cr.target,
            "ucl": cr.h * cr.sigma,
            "lcl": -cr.h * cr.sigma,
            "chart_data": cr.cusum_pos,
            "chart_data_neg": cr.cusum_neg,
            "in_control": cr.in_control,
            "out_of_control": [
                {"index": i, "value": cr.cusum_pos[i] if i in cr.signals_up else cr.cusum_neg[i], "rule": 1}
                for i in sorted(ooc_indices)
            ],
            "violations": [],
            "n": cr.n,
            "chart_type": "cusum",
        }
        _add_capability(result, vals, usl, lsl)
        return result
    elif chart_type == "ewma":
        er = ewma_chart(vals)
        ooc = set(er.out_of_control_indices)
        result = {
            "mean": er.target,
            "ucl": er.ucl_steady,
            "lcl": er.lcl_steady,
            "chart_data": er.ewma_values,
            "ucl_series": er.ucl,
            "lcl_series": er.lcl,
            "in_control": er.in_control,
            "out_of_control": [{"index": i, "value": er.ewma_values[i], "rule": 1} for i in sorted(ooc)],
            "violations": [],
            "n": er.n,
            "chart_type": "ewma",
        }
        _add_capability(result, vals, usl, lsl)
        return result
    else:
        # Default: I-MR
        r = individuals_moving_range_chart(vals, usl=usl, lsl=lsl)

    # ── Standard Shewhart result ────────────────────────────────
    ooc_list = []
    for p in r.out_of_control:
        ooc_list.append({"index": p.get("index", 0), "value": p.get("value", 0), "rule": p.get("rule", 1)})

    # Apply Nelson rules only if requested
    violations = []
    if rules_mode != "none":
        for v in r.run_violations:
            violations.append(
                {
                    "index": v.get("index", 0),
                    "value": v.get("value", None),
                    "rule": v.get("rule_number", 0),
                    "desc": v.get("description", ""),
                }
            )

    result = {
        "mean": r.limits.cl,
        "ucl": r.limits.ucl,
        "lcl": r.limits.lcl,
        "in_control": r.in_control if rules_mode != "none" else (len(ooc_list) == 0),
        "out_of_control": ooc_list,
        "violations": violations,
        "n": len(vals),
        "chart_type": chart_type,
    }

    # Secondary chart (MR for I-MR, R for X-bar/R, S for X-bar/S)
    if r.secondary_chart:
        sc = r.secondary_chart
        result["secondary"] = {
            "chart_type": sc.chart_type,
            "mean": sc.limits.cl,
            "ucl": sc.limits.ucl,
            "lcl": sc.limits.lcl,
            "data_points": sc.data_points,
        }

    _add_capability(result, vals, usl, lsl)
    return result


def _add_capability(result, vals, usl, lsl):
    """Append full capability indices if spec limits provided."""
    if usl is None and lsl is None:
        return
    from forgespc.capability import calculate_capability

    cap = calculate_capability(vals, usl=usl, lsl=lsl)
    result["cpk"] = cap.cpk
    result["cp"] = cap.cp
    result["pp"] = cap.pp
    result["ppk"] = cap.ppk
    result["dpmo"] = cap.dpmo
    result["sigma_level"] = cap.sigma_level
    result["yield_percent"] = cap.yield_percent
    result["ppm"] = round(cap.dpmo) if cap.dpmo is not None else None


@_rack_op("histogram")
def _op_histogram(d):
    vals = d["values"]
    n_bins = d.get("bins", 15)
    n = len(vals)

    if n_bins == 0 or n_bins is None:
        # Sturges' rule
        import math as _math

        n_bins = max(1, int(_math.ceil(_math.log2(n) + 1)))

    mn, mx = min(vals), max(vals)
    rng = mx - mn or 1
    bin_width = rng / n_bins
    bins = [0] * n_bins
    edges = [mn + i * bin_width for i in range(n_bins + 1)]

    for v in vals:
        idx = int((v - mn) / bin_width)
        if idx >= n_bins:
            idx = n_bins - 1
        bins[idx] += 1

    mean = sum(vals) / n
    variance = sum((v - mean) ** 2 for v in vals) / (n - 1) if n > 1 else 0
    std = variance**0.5

    # Skewness and kurtosis
    if n > 2 and std > 0:
        skew = (n / ((n - 1) * (n - 2))) * sum(((v - mean) / std) ** 3 for v in vals)
    else:
        skew = 0.0
    if n > 3 and std > 0:
        k4 = sum(((v - mean) / std) ** 4 for v in vals)
        kurt = ((n * (n + 1)) / ((n - 1) * (n - 2) * (n - 3))) * k4 - (3 * (n - 1) ** 2) / ((n - 2) * (n - 3))
    else:
        kurt = 0.0

    return {
        "bins": bins,
        "edges": edges,
        "n_bins": n_bins,
        "bin_width": bin_width,
        "mean": mean,
        "std": std,
        "skewness": skew,
        "kurtosis": kurt,
        "n": n,
        "min": mn,
        "max": mx,
    }


@_rack_op("gage_rr")
def _op_gage_rr(d):
    from forgespc.gage import gage_rr_crossed

    parts = d["parts"]
    operators = d["operators"]
    measurements = d["measurements"]

    # Validate balanced design before calling forgespc
    combos = set()
    for p, o in zip(parts, operators):
        combos.add((str(p), str(o)))
    unique_parts = set(str(p) for p in parts)
    unique_ops = set(str(o) for o in operators)
    expected = len(unique_parts) * len(unique_ops)
    actual_combos = len(set((str(p), str(o)) for p, o in zip(parts, operators)))
    if actual_combos < expected:
        missing = expected - actual_combos
        return {
            "error_type": "unbalanced_design",
            "message": (
                f"Unbalanced design: {len(unique_parts)} parts x {len(unique_ops)} operators = "
                f"{expected} combinations expected, but only {actual_combos} found ({missing} missing). "
                f"Each part must be measured by every operator for crossed Gage R&R."
            ),
            "n_parts": len(unique_parts),
            "n_operators": len(unique_ops),
            "parts": sorted(unique_parts),
            "operators": sorted(unique_ops),
        }

    result = gage_rr_crossed(
        parts=parts,
        operators=operators,
        measurements=measurements,
        tolerance=d.get("tolerance"),
    )
    return {
        "repeatability": result.var_repeatability,
        "reproducibility": result.var_reproducibility,
        "grr": result.var_grr,
        "part_variation": result.var_part,
        "total_variation": result.var_total,
        "grr_percent": result.grr_percent,
        "ndc": result.ndc,
        "assessment": result.assessment,
        "n_parts": result.n_parts,
        "n_operators": result.n_operators,
        "n_replicates": result.n_replicates,
        "n_total": result.n_total,
        "pct_contribution": result.pct_contribution,
        "pct_study_var": result.pct_study_var,
    }


@_rack_op("doe_design")
def _op_doe_design(d):
    from forgedoe.core.types import Factor, FactorType
    from forgedoe.designs.classical import latin_square, randomized_block, taguchi
    from forgedoe.designs.evop import evop_phase
    from forgedoe.designs.factorial import fractional_factorial, full_factorial, plackett_burman
    from forgedoe.designs.mixture import extreme_vertices, simplex_centroid, simplex_lattice
    from forgedoe.designs.optimal import d_optimal, i_optimal
    from forgedoe.designs.response_surface import box_behnken_design, central_composite_design
    from forgedoe.designs.screening import definitive_screening_design
    from forgedoe.designs.space_filling import latin_hypercube, maximin_lhs
    from forgedoe.designs.split_plot import split_plot, split_plot_ccd

    factors = []
    for f in d.get("factors", []):
        if f.get("factor_type") == "categorical":
            factors.append(Factor(f["name"], factor_type=FactorType.CATEGORICAL, levels=f.get("levels", [])))
        else:
            factors.append(Factor(f["name"], f.get("low", -1), f.get("high", 1)))
    design_type = d.get("design", "full")
    n_runs_req = d.get("n_runs", 20)

    # ── Design dispatch ─────────────────────────────────────────
    designers = {
        # Factorial
        "full": lambda: full_factorial(factors, randomize=True),
        "fractional": lambda: fractional_factorial(factors, resolution=d.get("resolution", 3), randomize=True),
        "plackett_burman": lambda: plackett_burman(factors, randomize=True),
        # Response surface
        "ccd": lambda: central_composite_design(factors, randomize=True),
        "bbd": lambda: box_behnken_design(factors, randomize=True),
        # Screening
        "dsd": lambda: definitive_screening_design(factors, randomize=True),
        # Classical
        "latin_square": lambda: latin_square([f.name for f in factors], randomize=True),
        "rcbd": lambda: randomized_block(len(factors), d.get("n_blocks", 3), d.get("replicates", 1)),
        "taguchi": lambda: taguchi(factors, array_type=d.get("array_type", "auto")),
        # Optimal
        "d_optimal": lambda: d_optimal(factors, n_runs_req, model=d.get("model", "linear")),
        "i_optimal": lambda: i_optimal(factors, n_runs_req, model=d.get("model", "quadratic")),
        # Space-filling
        "lhs": lambda: latin_hypercube(factors, n_samples=n_runs_req, randomize=True),
        "maximin_lhs": lambda: maximin_lhs(factors, n_samples=n_runs_req),
        # Mixture
        "simplex_lattice": lambda: simplex_lattice(len(factors), degree=d.get("degree", 2)),
        "simplex_centroid": lambda: simplex_centroid(len(factors), augment_axial=d.get("augment", False)),
        "extreme_vertices": lambda: extreme_vertices(len(factors)),
        # Split-plot
        "split_plot": lambda: split_plot(
            factors[: len(factors) // 2],
            factors[len(factors) // 2 :],
            n_replicates=d.get("replicates", 1),
        ),
        "split_plot_ccd": lambda: split_plot_ccd(
            factors[: len(factors) // 2],
            factors[len(factors) // 2 :],
        ),
        # EVOP
        "evop": lambda: evop_phase(factors),
    }

    if design_type not in designers:
        return {
            "error_type": "unknown_design",
            "message": f"Unknown design: {design_type}. Available: {sorted(designers.keys())}",
        }

    dm = designers[design_type]()

    # Mixture designs have no natural scale (proportions), others do
    try:
        nat = dm.to_natural()
        matrix = nat.matrix
    except Exception:
        matrix = dm.matrix

    # Build run sheet as list of dicts — map categorical coded values to level names
    runs = []
    for i in range(dm.n_runs):
        row = {"run": dm.run_order[i] if dm.run_order else i + 1}
        for j, f in enumerate(dm.factors):
            val = matrix[i][j]
            if f.factor_type == FactorType.CATEGORICAL and f.levels:
                # Map coded value to level: round to nearest int index
                idx = int(round((val + 1) / 2 * (len(f.levels) - 1)))
                idx = max(0, min(idx, len(f.levels) - 1))
                row[f.name] = f.levels[idx]
            else:
                row[f.name] = val
        runs.append(row)

    return {
        "design_type": dm.design_type,
        "n_runs": dm.n_runs,
        "n_factors": dm.n_factors,
        "factor_names": [f.name for f in dm.factors],
        "runs": runs,
        "coded_matrix": dm.matrix,
    }


@_rack_op("doe_analyze")
def _op_doe_analyze(d):
    from forgedoe.analysis.regression import fit_model

    coded_matrix = d.get("coded_matrix", [])
    responses = d.get("responses", [])
    factor_names = d.get("factor_names", [])
    model_type = d.get("model", "linear+interactions")
    alpha = float(d.get("alpha", 0.05))

    if not coded_matrix or not responses:
        return {"error_type": "missing_data", "message": "Need coded_matrix and responses."}
    if len(coded_matrix) != len(responses):
        return {
            "error_type": "length_mismatch",
            "message": f"Matrix has {len(coded_matrix)} runs but got {len(responses)} responses.",
        }

    ar = fit_model(coded_matrix, responses, factor_names, model_type=model_type, alpha=alpha)

    return {
        "model_type": ar.model_type,
        "coefficients": ar.coefficients,
        "se_coefficients": ar.se_coefficients,
        "t_values": ar.t_values,
        "p_values": ar.p_values,
        "significant_terms": ar.significant_terms,
        "r_squared": ar.r_squared,
        "r_squared_adj": ar.r_squared_adj,
        "residual_std": ar.residual_std,
        "f_statistic": ar.f_statistic,
        "f_p_value": ar.f_p_value,
        "effects": ar.effects,
        "alpha": ar.alpha,
    }


@_rack_op("doe_optimize")
def _op_doe_optimize(d):
    from forgedoe.analysis.optimization import optimize_responses
    from forgedoe.analysis.regression import fit_model
    from forgedoe.core.types import Factor, Response

    coded_matrix = d.get("coded_matrix", [])
    responses_data = d.get("responses", {})  # {name: {values: [], goal: "maximize|minimize|target", ...}}
    factor_defs = d.get("factors", [])
    factor_names = d.get("factor_names", [])
    model_type = d.get("model", "linear+interactions")
    alpha = float(d.get("alpha", 0.05))

    if not coded_matrix or not responses_data:
        return {"error_type": "missing_data", "message": "Need coded_matrix and responses."}

    factors = [Factor(f["name"], f["low"], f["high"]) for f in factor_defs]

    # Fit model for each response
    analysis_results = {}
    response_specs = []
    for rname, rspec in responses_data.items():
        vals = rspec.get("values", [])
        if len(vals) != len(coded_matrix):
            return {
                "error_type": "length_mismatch",
                "message": f"Response '{rname}': {len(vals)} values but {len(coded_matrix)} runs.",
            }
        ar = fit_model(coded_matrix, vals, factor_names, model_type=model_type, alpha=alpha)
        analysis_results[rname] = ar

        goal = rspec.get("goal", "maximize")
        response_specs.append(
            Response(
                name=rname,
                maximize=(goal == "maximize"),
                minimize=(goal == "minimize"),
                target=rspec.get("target"),
                lower_limit=rspec.get("lower_limit"),
                upper_limit=rspec.get("upper_limit"),
                importance=rspec.get("importance", 1.0),
            )
        )

    opt = optimize_responses(analysis_results, factors, response_specs, n_starts=d.get("n_starts", 200))

    return {
        "optimal_settings": opt.optimal_settings,
        "predicted_responses": opt.predicted_responses,
        "desirability": opt.desirability,
    }


@_rack_op("regression")
def _op_regression(d):
    x, y = d["x"], d["y"]
    n = len(x)
    mx = sum(x) / n
    my = sum(y) / n
    num = sum((x[i] - mx) * (y[i] - my) for i in range(n))
    den = sum((x[i] - mx) ** 2 for i in range(n))
    slope = num / den if den else 0
    intercept = my - slope * mx
    return {"slope": slope, "intercept": intercept, "n": n}


@require_auth
@require_http_methods(["POST"])
def rack_compute(request):
    """Dispatch statistical operations to forge packages.

    POST /api/rack/compute/
    Body: {"op": "ttest_2sample", "data": {"a": [...], "b": [...]}}
    Returns: {"result": {...}} or {"error": "..."}
    """
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    op = body.get("op")
    data = body.get("data")

    if not op or not data:
        return JsonResponse({"error": "Missing op or data"}, status=400)

    if op not in _RACK_OPS:
        return JsonResponse(
            {"error": f"Unknown op: {op}", "available": list(_RACK_OPS.keys())},
            status=400,
        )

    # Coerce string values to floats in any list fields
    import logging

    logger = logging.getLogger("forgerack.compute")

    # Coerce lists that look numeric (first non-null element is a number)
    for key, val in data.items():
        if isinstance(val, list) and len(val) > 0:
            # Check if the first non-null value is numeric
            first = next((v for v in val if v is not None and str(v).strip()), None)
            try:
                float(first)
                is_numeric = True
            except (TypeError, ValueError):
                is_numeric = False
            if is_numeric:
                coerced = []
                for v in val:
                    try:
                        coerced.append(float(v))
                    except (TypeError, ValueError):
                        pass
                data[key] = coerced

    logger.info(
        "rack compute op=%s keys=%s lens=%s",
        op,
        list(data.keys()),
        {k: len(v) if isinstance(v, list) else type(v).__name__ for k, v in data.items()},
    )
    if op in ("pearson", "spearman") and "x" in data and "y" in data:
        logger.info("  x[:5]=%s y[:5]=%s", data["x"][:5], data["y"][:5])

    try:
        result = _RACK_OPS[op](data)
        # Sanitize NaN/Inf → null for valid JSON
        import math

        def _clean(v):
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                logger.warning("NaN/Inf in %s result for op=%s", v, op)
                return 0.0
            return v

        result = {k: _clean(v) for k, v in result.items()}
        return JsonResponse({"result": result})
    except Exception as e:
        logger.exception("rack compute error: op=%s", op)
        return JsonResponse({"error": str(e)}, status=422)


@require_auth
@require_http_methods(["POST"])
def rack_export_runsheet(request):
    """Export a DOE run sheet as Excel (.xlsx).

    POST /api/rack/export-runsheet/
    Body: {"factor_names": ["Temp", "Press"], "runs": [{"run": 1, "Temp": 150, "Press": 2.0}, ...], "design_type": "Full Factorial"}
    Returns: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    factor_names = body.get("factor_names", [])
    runs = body.get("runs", [])
    design_type = body.get("design_type", "DOE")
    n_runs = body.get("n_runs", len(runs))

    if not factor_names or not runs:
        return JsonResponse({"error": "Missing factor_names or runs"}, status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Run Sheet"

    # Styles
    header_font = Font(name="Courier New", bold=True, size=10)
    header_fill = PatternFill(start_color="2A2E22", end_color="2A2E22", fill_type="solid")
    header_font_white = Font(name="Courier New", bold=True, size=10, color="A3B18A")
    data_font = Font(name="Courier New", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(factor_names) + 2)
    ws["A1"] = f"{design_type} — {n_runs} runs"
    ws["A1"].font = Font(name="Courier New", bold=True, size=12)

    # Headers: Run | Factor1 | Factor2 | ... | Response
    headers = ["Run"] + factor_names + ["Response"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, run in enumerate(runs, 4):
        # Run number
        cell = ws.cell(row=row_idx, column=1, value=run.get("run", row_idx - 3))
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # Factor values
        for col_idx, fname in enumerate(factor_names, 2):
            val = run.get(fname, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Empty response column
        cell = ws.cell(row=row_idx, column=len(factor_names) + 2, value="")
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 8
    for col_idx, fname in enumerate(factor_names, 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(fname) + 4, 12)
    ws.column_dimensions[get_column_letter(len(factor_names) + 2)].width = 14

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_name = design_type.replace(" ", "_").replace("/", "-")
    response["Content-Disposition"] = f'attachment; filename="runsheet_{safe_name}_{n_runs}runs.xlsx"'
    return response
